r"""Convert Kingdee purchase order Excel to the lightweight CSV used by the dashboard.

Default input:
    D:\BI文件\2026年销售出库汇总表\Fac-采购订单列表.xlsx

Default output:
    D:\BI文件\2026年销售出库汇总表\Fac-金蝶采购订单列表-轻量.csv
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path(r"D:\BI文件\2026年销售出库汇总表\Fac-采购订单列表.xlsx")
DEFAULT_OUTPUT = Path(r"D:\BI文件\2026年销售出库汇总表\Fac-金蝶采购订单列表-轻量.csv")
PREFERRED_SHEET = "Fac-采购订单列表"

OUTPUT_HEADERS = [
    "事业部",
    "单据编号",
    "供应商",
    "物料编码",
    "SKU",
    "物料名称",
    "采购数量",
    "剩余入库数量",
    "创建人",
]

ALIASES = {
    "事业部": ["事业部", "部门", "业务部门"],
    "单据编号": ["单据编号", "单据号", "采购订单号", "采购订单编号", "订单编号", "采购单号"],
    "供应商": ["供应商", "供应商名称", "供方", "厂家", "厂商"],
    "物料编码": ["物料编码", "品号", "物料代码", "商品编码", "存货编码", "产品编码"],
    "SKU": ["SKU", "sku", "领星SKU", "商品SKU", "物料SKU"],
    "物料名称": ["物料名称", "商品名称", "物品名称", "存货名称", "产品名称", "金蝶名称", "品名"],
    "采购数量": ["采购数量", "数量", "订单数量"],
    "剩余入库数量": ["剩余入库数量", "未入库数量", "剩余数量"],
    "创建人": ["创建人", "采购订单下单人", "下单人", "制单人"],
}


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", "", text.strip())
    text = text.replace("(", "").replace(")", "").replace("（", "").replace("）", "")
    return text.lower()


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_header_map(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        key: {normalize_header(alias) for alias in aliases}
        for key, aliases in ALIASES.items()
    }
    for row_index, row in enumerate(rows):
        headers = [normalize_header(cell) for cell in row]
        header_map: dict[str, int] = {}
        for key, alias_set in normalized_aliases.items():
            for column_index, header in enumerate(headers):
                if header in alias_set:
                    header_map[key] = column_index
                    break
        if "物料编码" in header_map and ("采购数量" in header_map or "剩余入库数量" in header_map):
            return row_index, header_map
    raise ValueError("未找到可识别表头，请确认源表包含物料编码、采购数量、剩余入库数量等字段")


def get_value(row: tuple[object, ...], header_map: dict[str, int], key: str) -> str:
    column_index = header_map.get(key)
    if column_index is None or column_index >= len(row):
        return ""
    return cell_text(row[column_index])


def convert(input_path: Path, output_path: Path) -> int:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet_name = next((name for name in workbook.sheetnames if PREFERRED_SHEET in name), workbook.sheetnames[0])
    worksheet = workbook[sheet_name]

    iterator = worksheet.iter_rows(values_only=True)
    preview_rows = []
    for _ in range(80):
        try:
            preview_rows.append(next(iterator))
        except StopIteration:
            break

    header_index, header_map = find_header_map(preview_rows)
    data_rows = preview_rows[header_index + 1 :]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()

        for row in data_rows:
            count += write_row(writer, row, header_map)
        for row in iterator:
            count += write_row(writer, row, header_map)

    return count


def write_row(writer: csv.DictWriter, row: tuple[object, ...], header_map: dict[str, int]) -> int:
    item = {header: get_value(row, header_map, header) for header in OUTPUT_HEADERS}
    if not any(item.values()):
        return 0
    if not (item["物料编码"] or item["SKU"] or item["物料名称"] or item["供应商"]):
        return 0
    writer.writerow(item)
    return 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert Kingdee purchase order Excel to lightweight dashboard CSV.")
    parser.add_argument("--input", "-i", type=Path, default=DEFAULT_INPUT, help="Source Kingdee purchase order .xlsx file.")
    parser.add_argument("--output", "-o", type=Path, default=DEFAULT_OUTPUT, help="Output lightweight .csv file.")
    args = parser.parse_args()

    count = convert(args.input, args.output)
    print(f"已生成轻量CSV：{args.output}")
    print(f"导出行数：{count}")


if __name__ == "__main__":
    main()
