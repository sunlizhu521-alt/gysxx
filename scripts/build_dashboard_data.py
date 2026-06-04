import base64
import json
import re
import tempfile
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SHARED_LIBRARY = DATA_DIR / "shared-library.json"


def text(value):
    return "" if value is None else str(value).strip()


def number(value):
    raw = text(value).replace(",", "")
    if not raw:
        return 0
    try:
        return float(raw)
    except ValueError:
        return 0


def normalize_material_code(value):
    return re.sub(r"\.0$", "", text(value)).lower()


def normalize_header(value):
    return re.sub(r"[\s()（）]", "", text(value)).lower()


def normalize_group_key(value):
    value = re.sub(r"[\s()（）【】\[\]_-]", "", text(value)).lower()
    if not value:
        return ""
    if "其他配件" in value or "其他/配件" in value:
        return "其他配件"
    if "一组" in value or "1组" in value:
        return "采购一组"
    if "二组" in value or "2组" in value:
        return "采购二组"
    if "三组" in value or "3组" in value:
        return "采购三组"
    if "四组" in value or "4组" in value:
        return "采购四组"
    return value


def format_region(address):
    value = re.sub(r"\s+", "", text(address))
    if not value:
        return ""
    direct = re.match(r"^(北京市|上海市|天津市|重庆市)([^省市自治区特别行政区]{1,12}?(区|县|市))?", value)
    if direct:
        return direct.group(1) + (direct.group(2) or "")
    province = re.match(r"^(.{2,12}?(省|自治区|特别行政区))", value)
    if not province:
        city_index = value.find("市")
        return value[: city_index + 1] if city_index > 0 else value[:12]
    rest = value[len(province.group(1)) :]
    city_index = rest.find("市")
    if city_index > 0:
        return province.group(1) + rest[: city_index + 1]
    region = re.match(r"^(.{2,12}?(自治州|地区|盟))", rest)
    return province.group(1) + (region.group(1) if region else "")


CITY_REGION_MAP = {
    "常州": "江苏省常州市",
    "泰兴": "江苏省泰州市",
    "苏州": "江苏省苏州市",
    "无锡": "江苏省无锡市",
    "宁波": "浙江省宁波市",
    "杭州": "浙江省杭州市",
    "嘉兴": "浙江省嘉兴市",
    "金华": "浙江省金华市",
    "台州": "浙江省台州市",
    "温州": "浙江省温州市",
    "中山": "广东省中山市",
    "广州": "广东省广州市",
    "佛山": "广东省佛山市",
    "东莞": "广东省东莞市",
    "深圳": "广东省深圳市",
    "厦门": "福建省厦门市",
    "上海": "上海市",
    "北京": "北京市",
    "天津": "天津市",
}


def infer_region_from_supplier_name(*names):
    text_value = "".join(text(name) for name in names)
    for city, region in CITY_REGION_MAP.items():
        if city in text_value:
            return region
    province_match = re.search(r"(江苏|浙江|广东|福建|上海|北京|天津|山东|安徽|河北|河南|湖北|湖南|江西|四川|重庆|陕西)", text_value)
    if province_match:
        province = province_match.group(1)
        return province if province.endswith("市") else f"{province}省"
    return ""


def is_stock_age_over_60(value):
    value = text(value)
    if not value:
        return False
    if re.search(r"库龄\s*[>＞]\s*60", value):
        return True
    if re.match(r"^(是|true|yes|y|1|超60|超过60天)$", value, re.I):
        return True
    return number(value) > 60


def load_shared_records():
    with SHARED_LIBRARY.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def workbook_from_record(record):
    suffix = Path(record["name"]).suffix or ".xlsx"
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp.write(base64.b64decode(record.get("dataBase64", "")))
    temp.close()
    return openpyxl.load_workbook(temp.name, read_only=True, data_only=True), Path(temp.name)


def find_record(payload, store_name, record_id):
    for record in payload.get("stores", {}).get(store_name, []):
        if record.get("id") == record_id:
            return record
    raise KeyError(f"Missing {store_name}/{record_id} in shared-library.json")


def rows_from_sheet(ws):
    return list(ws.iter_rows(values_only=True))


def build_category_map(category_record):
    wb, temp_path = workbook_from_record(category_record)
    try:
        sheet_name = "Dim-YL医疗器械商品分类" if "Dim-YL医疗器械商品分类" in wb.sheetnames else wb.sheetnames[0]
        rows = rows_from_sheet(wb[sheet_name])
        category_map = {}
        group_map = {}
        for row in rows[1:]:
            material_code = normalize_material_code(row[0] if len(row) > 0 else "")
            if not material_code:
                continue
            item = {
                "sku": text(row[2] if len(row) > 2 else ""),
                "materialName": text(row[3] if len(row) > 3 else ""),
                "salesLine": text(row[6] if len(row) > 6 else ""),
                "salesSeries": text(row[7] if len(row) > 7 else ""),
                "purchaseGroup": text(row[20] if len(row) > 20 else ""),
            }
            category_map[material_code] = item
            group_key = normalize_group_key(item["purchaseGroup"])
            if group_key:
                group_map[group_key] = item
        return category_map, group_map
    finally:
        wb.close()
        temp_path.unlink(missing_ok=True)


def find_sheet_name(workbook, preferred, fallback_contains):
    if preferred in workbook.sheetnames:
        return preferred
    for name in workbook.sheetnames:
        if preferred and preferred in name:
            return name
    for name in workbook.sheetnames:
        if fallback_contains in name:
            return name
    return workbook.sheetnames[0]


def build_supplier_directory(payload, category_map, group_map):
    purchase_record = find_record(payload, "dimension-files", "dimension-6")
    wb, temp_path = workbook_from_record(purchase_record)
    try:
        detail_sheet = find_sheet_name(wb, "产品线明细", "明细")
        rows = rows_from_sheet(wb[detail_sheet])
        headers = [text(cell) for cell in rows[0]]
        records = []
        for index, row in enumerate(rows[1:]):
            material_code = text(row[3] if len(row) > 3 else "")
            supplier = text(row[6] if len(row) > 6 else "")
            supplier_short = text(row[7] if len(row) > 7 else "") or supplier
            if not any([material_code, supplier, supplier_short]):
                continue
            matched = category_map.get(normalize_material_code(material_code), {})
            region = format_region(row[19] if len(row) > 19 else "") or infer_region_from_supplier_name(supplier, supplier_short) or "未维护地址"
            records.append(
                {
                    "id": f"{index}-{material_code}-{supplier_short}",
                    "primaryLine": matched.get("salesLine") or "未匹配",
                    "secondaryLine": text(row[1] if len(row) > 1 else ""),
                    "group": matched.get("purchaseGroup") or "未匹配",
                    "owner": text(row[2] if len(row) > 2 else "") or "未分配",
                    "materialCode": material_code,
                    "sku": matched.get("sku") or "",
                    "materialName": matched.get("materialName") or "",
                    "supplier": supplier,
                    "supplierShort": supplier_short,
                    "purchasePrice": 0,
                    "unitPrice": 0,
                    "moq": number(row[8] if len(row) > 8 else ""),
                    "leadTime": number(row[9] if len(row) > 9 else ""),
                    "paymentTerm": text(row[13] if len(row) > 13 else "") or "未填写",
                    "contact": text(row[17] if len(row) > 17 else ""),
                    "phone": text(row[18] if len(row) > 18 else ""),
                    "address": text(row[19] if len(row) > 19 else ""),
                    "dimProductLine": matched.get("salesLine") or "",
                    "dimPurchaseGroup": matched.get("purchaseGroup") or "",
                    "region": region,
                }
            )

        division_rows = []
        division_headers = []
        division_sheet = next((name for name in wb.sheetnames if "产品线分工表" in name), None)
        if division_sheet:
            division_raw = rows_from_sheet(wb[division_sheet])
            header_index = next((idx for idx, row in enumerate(division_raw) if any(text(cell) for cell in row)), None)
            if header_index is not None:
                division_headers = [text(cell) for cell in division_raw[header_index]]
                group_index = next((idx for idx, header in enumerate(division_headers) if normalize_header(header) == normalize_header("组名")), 0)
                for row in division_raw[header_index + 1 :]:
                    if not any(text(cell) for cell in row):
                        continue
                    cells = [text(row[idx] if idx < len(row) else "") for idx in range(len(division_headers))]
                    group_name = cells[group_index] if group_index < len(cells) else ""
                    group_key = normalize_group_key(group_name)
                    division_rows.append(
                        {
                            "key": cells[0] if cells else "",
                            "cells": cells,
                            "groupName": group_name,
                            "groupKey": group_key,
                            "matchedPurchaseGroup": group_map.get(group_key, {}).get("purchaseGroup") or group_name,
                        }
                    )

        return {
            "version": 1,
            "generatedAt": now_iso(),
            "source": {
                "purchaseAssignment": pick_manifest_fields(purchase_record),
                "categoryDimension": pick_manifest_fields(find_record(payload, "dimension-files", "dimension-1")),
            },
            "records": records,
            "divisionRows": division_rows,
            "divisionHeaders": division_headers,
            "sourceHeaders": headers,
        }
    finally:
        wb.close()
        temp_path.unlink(missing_ok=True)


ORDER_ALIASES = {
    "materialCode": ["物料编码", "商品编码", "存货编码", "产品编码", "品号"],
    "sku": ["SKU", "sku", "领星SKU"],
    "itemName": ["物品名称", "物料名称", "商品名称", "存货名称", "产品名称", "金蝶名称", "品名"],
    "orderedQty": ["下单数量-备货需求-OA申请为准"],
    "shippedQty": ["发货数量"],
    "remainingQty": ["未发货数量"],
    "stockAge": ["库龄＞60", "库龄>60", "库龄"],
}


def create_header_map(headers, aliases):
    header_map = {}
    normalized_headers = [normalize_header(header) for header in headers]
    for key, names in aliases.items():
        normalized_names = [normalize_header(name) for name in names]
        for index, header in enumerate(normalized_headers):
            if header in normalized_names:
                header_map[key] = index
                break
    return header_map


def material_code_column_score(value):
    raw = text(value)
    if not raw:
        return 0
    normalized = normalize_material_code(raw)
    if len(normalized) < 4 or re.search(r"[\u4e00-\u9fff]", normalized):
        return 0
    score = 1
    if re.search(r"\d", normalized):
        score += 2
    if re.fullmatch(r"[a-z0-9._\-]+", normalized):
        score += 1
    return score


def infer_material_code_column(headers, data_rows, used_columns):
    unused_blank_columns = [
        index
        for index, header in enumerate(headers)
        if index not in used_columns and not text(header)
    ]
    candidate_columns = unused_blank_columns or [
        index for index in range(len(headers)) if index not in used_columns
    ]
    best_column = None
    best_score = 0
    for column in candidate_columns:
        score = sum(
            material_code_column_score(row[column] if column < len(row) else "")
            for row in data_rows[:50]
        )
        if score > best_score:
            best_column = column
            best_score = score
    return best_column if best_score >= 3 else None


def parse_order_rows(rows, business_unit):
    header_index = None
    header_map = None
    for index, row in enumerate(rows):
        headers = [text(cell) for cell in row]
        possible_map = create_header_map(headers, ORDER_ALIASES)
        if all(key in possible_map for key in ["orderedQty", "shippedQty", "remainingQty"]):
            if "materialCode" not in possible_map:
                inferred_column = infer_material_code_column(
                    headers,
                    rows[index + 1 :],
                    set(possible_map.values()),
                )
                if inferred_column is not None:
                    possible_map["materialCode"] = inferred_column
        if all(key in possible_map for key in ["materialCode", "orderedQty", "shippedQty", "remainingQty"]):
            header_index = index
            header_map = possible_map
            break
    if header_index is None:
        return []

    records = []
    for index, row in enumerate(rows[header_index + 1 :]):
        def cell(key):
            col = header_map.get(key)
            return text(row[col] if col is not None and col < len(row) else "")

        material_code = cell("materialCode")
        sku = cell("sku")
        item_name = cell("itemName")
        if not any([material_code, sku, item_name]):
            continue
        ordered_qty = number(cell("orderedQty"))
        shipped_qty = number(cell("shippedQty"))
        remaining_qty = number(cell("remainingQty"))
        records.append(
            {
                "id": f"{business_unit}-{index}-{material_code}-{sku}",
                "businessUnit": business_unit,
                "materialCode": material_code,
                "sku": sku,
                "itemName": item_name,
                "orderedQty": ordered_qty,
                "shippedQty": shipped_qty,
                "undeliveredQty": remaining_qty,
                "remainingQty": remaining_qty,
                "isOver60": is_stock_age_over_60(cell("stockAge")),
            }
        )
    return records


def business_unit_from_sheet_name(sheet_name):
    name = text(sheet_name)
    return re.sub(r"[（(].*?[）)]", "", name).strip() or name or "未匹配"


def build_delivery_orders(payload, category_map):
    fact_record = find_record(payload, "fact-files", "fact-1")
    wb, temp_path = workbook_from_record(fact_record)
    try:
        records = []
        for sheet_name in wb.sheetnames:
            rows = rows_from_sheet(wb[sheet_name])
            for record in parse_order_rows(rows, business_unit_from_sheet_name(sheet_name)):
                matched = category_map.get(normalize_material_code(record["materialCode"]), {})
                record["salesLine"] = matched.get("salesLine") or "未匹配"
                record["salesSeries"] = matched.get("salesSeries") or "未匹配"
                record["purchaseGroup"] = matched.get("purchaseGroup") or "未匹配"
                records.append(record)
        return {
            "version": 1,
            "generatedAt": now_iso(),
            "source": {
                "purchaseOrder": pick_manifest_fields(fact_record),
                "categoryDimension": pick_manifest_fields(find_record(payload, "dimension-files", "dimension-1")),
            },
            "records": records,
        }
    finally:
        wb.close()
        temp_path.unlink(missing_ok=True)


def pick_manifest_fields(record):
    return {
        "id": record.get("id"),
        "name": record.get("name"),
        "size": record.get("size"),
        "refreshMonth": record.get("refreshMonth"),
        "savedAt": record.get("savedAt"),
        "appliedAt": record.get("appliedAt"),
    }


def now_iso():
    tz = timezone(timedelta(hours=8))
    return datetime.now(tz).replace(microsecond=0).isoformat()


def build_manifest(payload, supplier_payload, delivery_payload):
    return {
        "version": 1,
        "generatedAt": now_iso(),
        "sharedLibraryGeneratedAt": payload.get("generatedAt"),
        "dashboards": {
            "supplierDirectory": {
                "file": "supplier-directory.json",
                "recordCount": len(supplier_payload["records"]),
                "source": supplier_payload["source"],
            },
            "deliveryOrders": {
                "file": "delivery-orders.json",
                "recordCount": len(delivery_payload["records"]),
                "source": delivery_payload["source"],
            },
        },
    }


def write_json(path, payload):
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))


def main():
    payload = load_shared_records()
    category_record = find_record(payload, "dimension-files", "dimension-1")
    category_map, group_map = build_category_map(category_record)
    supplier_payload = build_supplier_directory(payload, category_map, group_map)
    delivery_payload = build_delivery_orders(payload, category_map)
    manifest_payload = build_manifest(payload, supplier_payload, delivery_payload)

    write_json(DATA_DIR / "supplier-directory.json", supplier_payload)
    write_json(DATA_DIR / "delivery-orders.json", delivery_payload)
    write_json(DATA_DIR / "library-manifest.json", manifest_payload)

    print(f"supplier-directory rows: {len(supplier_payload['records'])}")
    print(f"delivery-orders rows: {len(delivery_payload['records'])}")


if __name__ == "__main__":
    main()
